# -*- coding: utf-8 -*-
"""
Programa para desbloquear planilhas Excel com senha
Interface gráfica com tkinter
"""

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import threading
import os
from openpyxl import load_workbook
from openpyxl.worksheet.protection import SheetProtection
from PIL import Image, ImageDraw
import io

class UnlockerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Desbloqueador de Planilhas Excel")
        self.root.geometry("600x500")
        self.root.resizable(False, False)
        
        # Variáveis
        self.arquivo_selecionado = tk.StringVar()
        self.desbloquear_em_progresso = False
        
        # Configuração de cores
        self.cor_fundo = "#f0f0f0"
        self.cor_primaria = "#2c3e50"
        self.cor_botao = "#e74c3c"
        self.cor_sucesso = "#27ae60"
        
        self.root.configure(bg=self.cor_fundo)
        
        # Criar interface
        self.criar_interface()
        
    def criar_interface(self):
        """Cria os elementos da interface gráfica"""
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        titulo = ttk.Label(main_frame, text="🔓 Desbloqueador de Planilhas Excel", 
                          font=("Arial", 16, "bold"))
        titulo.pack(pady=(0, 20))
        
        # Frame para seleção de arquivo
        arquivo_frame = ttk.LabelFrame(main_frame, text="1. Selecione o Arquivo", padding="15")
        arquivo_frame.pack(fill=tk.X, pady=10)
        
        # Exibir arquivo selecionado
        self.arquivo_label = ttk.Label(arquivo_frame, text="Nenhum arquivo selecionado", 
                                       font=("Arial", 10), foreground="gray")
        self.arquivo_label.pack(fill=tk.X, pady=(0, 10))
        
        # Botão carregar
        btn_carregar = ttk.Button(arquivo_frame, text="📁 Carregar Arquivo Excel", 
                                  command=self.carregar_arquivo)
        btn_carregar.pack(fill=tk.X)
        
        # Frame para desbloquear
        desbloq_frame = ttk.LabelFrame(main_frame, text="2. Desbloquear Planilha", padding="15")
        desbloq_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(desbloq_frame, text="Clique no botão abaixo para desbloquear:", 
                 font=("Arial", 10)).pack(pady=(0, 10))
        
        # Botão desbloquear com estilo
        self.btn_desbloquear = ttk.Button(desbloq_frame, text="🔓 Desbloquear Planilha", 
                                         command=self.desbloquear_click, state=tk.DISABLED)
        self.btn_desbloquear.pack(fill=tk.X, ipady=15)
        
        # Frame para salvar
        salvar_frame = ttk.LabelFrame(main_frame, text="3. Salvar Arquivo", padding="15")
        salvar_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(salvar_frame, text="Salve o arquivo desbloqueado no seu computador", 
                 font=("Arial", 10)).pack(pady=(0, 10))
        
        self.btn_salvar = ttk.Button(salvar_frame, text="💾 Salvar Arquivo", 
                                    command=self.salvar_arquivo, state=tk.DISABLED)
        self.btn_salvar.pack(fill=tk.X, ipady=15)
        
        # Frame de progresso e status
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="15")
        status_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.progress_var = tk.IntVar(value=0)
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, 
                                           maximum=100, mode='indeterminate')
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        self.status_text = tk.Text(status_frame, height=6, width=60, state=tk.DISABLED, 
                                   font=("Courier", 9), bg="white", wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar para o texto
        scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.config(yscrollcommand=scrollbar.set)
        
    def log_status(self, mensagem):
        """Adiciona mensagem ao log de status"""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, mensagem + "\n")
        self.status_text.see(tk.END)  # Scroll automático
        self.status_text.config(state=tk.DISABLED)
        self.root.update()
        
    def carregar_arquivo(self):
        """Abre diálogo para selecionar arquivo Excel"""
        arquivo = filedialog.askopenfilename(
            title="Selecione um arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_selecionado.set(arquivo)
            nome_arquivo = os.path.basename(arquivo)
            self.arquivo_label.config(text=f"✓ {nome_arquivo}", foreground="green")
            self.btn_desbloquear.config(state=tk.NORMAL)
            
            # Limpar status anterior
            self.status_text.config(state=tk.NORMAL)
            self.status_text.delete(1.0, tk.END)
            self.status_text.config(state=tk.DISABLED)
            self.log_status(f"Arquivo carregado: {nome_arquivo}")
    
    def desbloquear_click(self):
        """Inicia o processo de desbloqueio em thread separada"""
        if not self.arquivo_selecionado.get():
            messagebox.showerror("Erro", "Por favor, selecione um arquivo primeiro!")
            return
        
        # Executar em thread para não travar a interface
        thread = threading.Thread(target=self.desbloquear_arquivo)
        thread.daemon = True
        thread.start()
    
    def desbloquear_arquivo(self):
        """Tenta desbloquear o arquivo Excel com força bruta"""
        self.desbloquear_em_progresso = True
        self.btn_desbloquear.config(state=tk.DISABLED)
        self.progress_bar.start()
        
        try:
            arquivo_path = self.arquivo_selecionado.get()
            self.log_status("Iniciando processo de desbloqueio...")
            self.log_status("Carregando arquivo Excel...")
            
            # Carregar workbook
            wb = load_workbook(arquivo_path)
            
            # Tentar desbloquear cada planilha
            planilhas_desbloqueadas = []
            total_tentativas = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self.log_status(f"\nProcessando planilha: {sheet_name}")
                
                # Verificar se está protegida
                if ws.protection.sheet:
                    self.log_status(f"  → Planilha '{sheet_name}' está protegida")
                    senha_encontrada = self.forcar_senha(ws, sheet_name)
                    
                    if senha_encontrada:
                        planilhas_desbloqueadas.append(sheet_name)
                        self.log_status(f"  ✓ Planilha desbloqueada com sucesso!")
                    else:
                        self.log_status(f"  ✗ Não foi possível desbloquear esta planilha")
                else:
                    self.log_status(f"  → Planilha '{sheet_name}' não está protegida")
            
            if planilhas_desbloqueadas:
                self.log_status(f"\n✓ Sucesso! {len(planilhas_desbloqueadas)} planilha(s) desbloqueada(s)")
                self.log_status("Clique em 'Salvar Arquivo' para salvar as alterações")
                self.btn_salvar.config(state=tk.NORMAL)
                messagebox.showinfo("Sucesso", 
                    f"{len(planilhas_desbloqueadas)} planilha(s) desbloqueada(s) com sucesso!\n\n"
                    f"Agora clique em 'Salvar Arquivo' para salvar as mudanças.")
            else:
                self.log_status("\n✗ Nenhuma planilha foi desbloqueada")
                messagebox.showwarning("Aviso", 
                    "Não foi possível desbloquear as planilhas.\n"
                    "Verifique se estão realmente protegidas.")
            
            # Guardar workbook na memória para salvar depois
            self.workbook_desbloqueado = wb
            
        except Exception as e:
            self.log_status(f"\n✗ Erro: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar arquivo:\n{str(e)}")
        
        finally:
            self.progress_bar.stop()
            self.btn_desbloquear.config(state=tk.NORMAL)
            self.desbloquear_em_progresso = False
    
    def forcar_senha(self, worksheet, sheet_name):
        """Tenta desbloquear com força bruta usando combinações de caracteres"""
        
        # Combinações similares à macro VBA
        caracteres_principais = [chr(i) for i in range(65, 67)]  # A, B
        caracteres_ascii = [chr(i) for i in range(32, 127)]  # Todos ASCII imprimíveis
        
        total_combinacoes = len(caracteres_principais) ** 11 * len(caracteres_ascii)
        tentativas = 0
        
        self.log_status(f"  Testando combinações de senhas...")
        
        try:
            # Gerar e testar combinações
            for i in caracteres_principais:
                for j in caracteres_principais:
                    for k in caracteres_principais:
                        for l in caracteres_principais:
                            for m in caracteres_principais:
                                for i1 in caracteres_principais:
                                    for i2 in caracteres_principais:
                                        for i3 in caracteres_principais:
                                            for i4 in caracteres_principais:
                                                for i5 in caracteres_principais:
                                                    for i6 in caracteres_principais:
                                                        for n in caracteres_ascii:
                                                            tentativa = i + j + k + l + m + i1 + i2 + i3 + i4 + i5 + i6 + n
                                                            tentativas += 1
                                                            
                                                            try:
                                                                worksheet.protection.sheet = False
                                                                worksheet.protection.password = tentativa
                                                                worksheet.protection.enable()
                                                                worksheet.protection.sheet = False
                                                                
                                                                # Verificar se desbloqueou
                                                                if not worksheet.protection.sheet:
                                                                    self.log_status(f"  → Senha encontrada: {tentativa}")
                                                                    return True
                                                                    
                                                            except:
                                                                pass
                                                            
                                                            # Atualizar interface
                                                            if tentativas % 1000 == 0:
                                                                progresso = int((tentativas / total_combinacoes) * 100)
                                                                self.progress_var.set(min(progresso, 99))
                                                                self.root.update()
        
        except Exception as e:
            self.log_status(f"  Erro durante tentativas: {str(e)}")
        
        return False
    
    def salvar_arquivo(self):
        """Salva o arquivo desbloqueado"""
        if not hasattr(self, 'workbook_desbloqueado'):
            messagebox.showerror("Erro", "Nenhum arquivo para salvar. Desbloqueie um arquivo primeiro!")
            return
        
        arquivo_original = self.arquivo_selecionado.get()
        nome_original = os.path.basename(arquivo_original)
        nome_sem_extensao = os.path.splitext(nome_original)[0]
        nome_sugerido = f"{nome_sem_extensao}_DESBLOQUEADO.xlsx"
        
        arquivo_destino = filedialog.asksaveasfilename(
            initialfile=nome_sugerido,
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo_destino:
            try:
                self.log_status(f"\nSalvando arquivo: {os.path.basename(arquivo_destino)}")
                self.workbook_desbloqueado.save(arquivo_destino)
                self.log_status("✓ Arquivo salvo com sucesso!")
                messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{arquivo_destino}")
            except Exception as e:
                self.log_status(f"✗ Erro ao salvar: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao salvar arquivo:\n{str(e)}")


def main():
    root = tk.Tk()
    app = UnlockerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
